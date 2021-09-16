from flask import Flask,request,jsonify,g,render_template
from flask import url_for, send_from_directory
import json
import os
from werkzeug.utils import secure_filename
import sqlite3
import pandas as pd
import numpy as np
from sklearn.model_selection import train_test_split
from eunjeon import Mecab
from collections import Counter
import tempfile
from tensorflow.keras.models import load_model
from tensorflow.keras.preprocessing.sequence import pad_sequences
from tensorflow.keras.preprocessing.text import Tokenizer
#from werkzeug import redirect
from flask.helpers import flash
from flask import send_file
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
import xlsxwriter
from datetime import datetime
from datetime import timedelta
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl import load_workbook
import openpyxl
from PIL import Image as PILImage
import xlrd
from collections import OrderedDict
import json
import time
import re


total_data = pd.read_table('ratings_total.txt', names=['ratings', 'reviews'])
total_data['label'] = np.select([total_data.ratings > 3], [1], default=0)
total_data['ratings'].nunique(), total_data['reviews'].nunique(), total_data['label'].nunique()
total_data.drop_duplicates(subset=['reviews'], inplace=True)

train_data, test_data = train_test_split(total_data, test_size = 0.25, random_state = 42)
train_data['reviews'] = train_data['reviews'].str.replace("[^ㄱ-ㅎㅏ-ㅣ가-힣 ]","")
train_data['reviews'].replace('', np.nan, inplace=True)
test_data.drop_duplicates(subset = ['reviews'], inplace=True) # 중복 제거
test_data['reviews'] = test_data['reviews'].str.replace("[^ㄱ-ㅎㅏ-ㅣ가-힣 ]","") # 정규 표현식 수행
test_data['reviews'].replace('', np.nan, inplace=True) # 공백은 Null 값으로 변경
test_data = test_data.dropna(how='any') # Null 값 제거
mecab = Mecab()

stopwords = ['도', '는', '다', '의', '가', '이', '은', '한', '에', '하', '고', '을', '를', '인', '듯', '과', '와', '네', '들', '듯', '지', '임', '게']
train_data['tokenized'] = train_data['reviews'].apply(mecab.morphs)
train_data['tokenized'] = train_data['tokenized'].apply(lambda x: [item for item in x if item not in stopwords])
test_data['tokenized'] = test_data['reviews'].apply(mecab.morphs)
test_data['tokenized'] = test_data['tokenized'].apply(lambda x: [item for item in x if item not in stopwords])
max_len = 80
negative_words = np.hstack(train_data[train_data.label == 0]['tokenized'].values)
positive_words = np.hstack(train_data[train_data.label == 1]['tokenized'].values)
negative_word_count = Counter(negative_words)
positive_word_count = Counter(positive_words)
text_len = train_data[train_data['label']==1]['tokenized'].map(lambda x: len(x))
text_len = train_data[train_data['label']==0]['tokenized'].map(lambda x: len(x))
X_train = train_data['tokenized'].values
y_train = train_data['label'].values
X_test= test_data['tokenized'].values
y_test = test_data['label'].values
tokenizer = Tokenizer()
tokenizer.fit_on_texts(X_train)

threshold = 2
total_cnt = len(tokenizer.word_index) # 단어의 수
rare_cnt = 0 # 등장 빈도수가 threshold보다 작은 단어의 개수를 카운트
total_freq = 0 # 훈련 데이터의 전체 단어 빈도수 총 합
rare_freq = 0 # 등장 빈도수가 threshold보다 작은 단어의 등장 빈도수의 총 합

# 단어와 빈도수의 쌍(pair)을 key와 value로 받는다.
for key, value in tokenizer.word_counts.items():
    total_freq = total_freq + value

    # 단어의 등장 빈도수가 threshold보다 작으면
    if(value < threshold):
        rare_cnt = rare_cnt + 1
        rare_freq = rare_freq + value

vocab_size = total_cnt - rare_cnt + 2

tokenizer = Tokenizer(vocab_size, oov_token = 'OOV') 
tokenizer.fit_on_texts(X_train)
X_train = tokenizer.texts_to_sequences(X_train)
X_test = tokenizer.texts_to_sequences(X_test)

X_train = pad_sequences(X_train, maxlen = max_len)
X_test = pad_sequences(X_test, maxlen = max_len)


loaded_model = load_model('best_model.h5')
loaded_model.compile(optimizer='rmsprop', loss='binary_crossentropy', metrics=['acc'])


#from pymongo import MongoClient
def json_trans(jsondata):
    strjson = str(jsondata).replace("[","")
    strjson = strjson.replace("]","") 
    return strjson

def sentiment_predict(new_sentence):
    new_sentence = mecab.morphs(new_sentence) # 토큰화
    new_sentence = [word for word in new_sentence if not word in stopwords] # 불용어 제거
    encoded = tokenizer.texts_to_sequences([new_sentence]) # 정수 인코딩
    pad_new = pad_sequences(encoded, maxlen = max_len) # 패딩
    score = float(loaded_model.predict(pad_new)) # 예측
    # if(score > 0.5):
    #     print("{:.2f}% 확률로 긍정 리뷰입니다.".format(score * 100))
    # else:
    #     print("{:.2f}% 확률로 부정 리뷰입니다.".format((1 - score) * 100))
    return score

def productSearch(productname):
    options = Options()
    options.add_argument('headless'); # headless는 화면이나 페이지 이동을 표시하지 않고 동작하는 모드
    
    workbook = xlsxwriter.Workbook("C:/Users/real1/OneDrive/Desktop/Tempus_flask/xlsx/"+productname+'.xlsx')
    worksheet = workbook.add_worksheet()
    excel_row = 1

    worksheet.set_column('A:A', 40) # A 열의 너비를 40으로 설정
    worksheet.set_row(0,18) # A열의 높이를 18로 설정
    worksheet.set_column('B:B', 12) # B 열의 너비를 12로 설정
    worksheet.set_column('C:C', 60) # C 열의 너비를 60으로 설정

    worksheet.write(0, 0, '제품 모델명')
    worksheet.write(0, 1, '가격')
    worksheet.write(0, 2, '링크')
    worksheet.write(0, 3, '리뷰')
    worksheet.write(0, 4, '감정분석 스코어')
    worksheet.write(0, 5, '추천점수')
    # webdirver 설정(Chrome, Firefox 등)
    # chromedriver_autoinstaller.install()
    driver = webdriver.Chrome(options=options) # 브라우저 창 안보이기
    # driver = webdriver.Chrome() # 브라우저 창 보이기
    
    # 크롬 브라우저 내부 대기 (암묵적 대기)
    driver.implicitly_wait(5)
    
    # 브라우저 사이즈
    driver.set_window_size(1920,1280)

    driver.get('https://shopping.naver.com/home/p/index.naver')# 네이버 쇼핑 메인창
    xpath2 = "/html/body/div[1]/div[1]/div/div[2]/div/div[2]/div/div[1]/form/fieldset/div[1]/input[1]"
    search = driver.find_element_by_name("query")
    search.clear()
    search.send_keys(productname)
    time.sleep(1)
    search.send_keys("\n")
    time.sleep(2)
    
    # 현재 페이지
    curPage = 1
    
    # 크롤링할 전체 페이지수
    totalPage = 1
    
    while curPage <= totalPage:
        #bs4 초기화
        soup = BeautifulSoup(driver.page_source, 'html.parser')
    
        # 상품 리스트 선택
        goods_list = soup.select('li.basicList_item__2XT81')
    
        # 페이지 번호 출력
        print('----- Current Page : {}'.format(curPage), '------')
    
        for v in goods_list:
            # 상품명, 가격, 이미지
            
            name = v.select_one('div.basicList_title__3P9Q7 > a').text.strip()
            price = v.select_one('span.price_num__2WUXn').text.strip()
            review_link = v.select_one('div.basicList_title__3P9Q7 > a').get('href')
            price = str(price).replace(',','')
            price = price.replace('원','') 
            print(name,', ', price, ', ', review_link)
            worksheet.write(excel_row, 0, name)
            worksheet.write(excel_row, 1, int(price))
            worksheet.write(excel_row, 2, review_link)



            driver.get(review_link)
            soup = BeautifulSoup(driver.page_source, 'html.parser')
            review_list = soup.select('p.reviewItems_text__XIsTc')
            temp_str = ""
            for r in review_list:

                r.text.strip()
                r = str(r)
                r.replace('<br/>','')
                r = re.sub('<.+?>','',r,0).strip()
                print(r)
                temp_str = temp_str + r +"\n"
                worksheet.write(excel_row,3,temp_str)
                print()
            excel_row += 1
            driver.back()
        
        
        time.sleep(5)
        print()
        # 페이지 수 증가
        curPage += 1
    
        if curPage > totalPage:
            print('Crawling succeed!')
            break
    
        # 페이지 이동 클릭
        cur_css = 'a.pagination_next__1ITTf'
        WebDriverWait(driver,3).until(EC.presence_of_element_located((By.CSS_SELECTOR,cur_css))).click()
    
        #2,3페이지까지 확인할경우 과한 리소스 소모+추천 제품에서 멀어져 정확도 감소-> 1페이지만 검색함

        # BeautifulSoup 인스턴스 삭제
        del soup
    
        # 3초간 대기
        time.sleep(3)
    
    # 브라우저 종료
    workbook.close()
    driver.close()    

def prdctRcmnd(prdctname):
    productSearch(prdctname)
    excel_path =  "C:/Users/real1/OneDrive/Desktop/Tempus_flask/xlsx/"+prdctname+".xlsx"
    wb = openpyxl.load_workbook(excel_path)
    sh = wb.active
    last_row = sh.max_row
    
    for rownum in range(1,last_row):
        score = sentiment_predict(str(sh["D"+str(rownum+1)].value))
        sh["E"+str(rownum+1)] = score
        sh["F"+str(rownum+1)] = (1/int(sh["B"+str(rownum+1)].value) * score) * 100000000
        print(score)
    wb.save(excel_path)
    wb.close()

    data = pd.read_excel("C:/Users/real1/OneDrive/Desktop/Tempus_flask/xlsx/"+prdctname+".xlsx", engine="openpyxl",thousands = ',')
    data = data.sort_values(by="추천점수",ascending=False)
    with pd.ExcelWriter("C:/Users/real1/OneDrive/Desktop/Tempus_flask/xlsx/"+prdctname+"_sort.xlsx",engine="openpyxl") as writer:
        data.to_excel(writer,sheet_name="sheet1",index=False) 




app = Flask(__name__)



@app.route('/')
def home():
    return "hello"
    # return render_template('C:/Users/real1/OneDrive/Desktop/Tempus_flask/upload.html')

# @app.route('/addboard',methods=['HEAD','GET','POST'])
# def addpost():
#     #print(request.is_json)
#     jsonData = request.get_json()
#     print(jsonData)
#     #print(jsonData[0]['Title'])
#     f = open("board.txt",'a')
#     strjson = str(jsonData).replace("[","")
#     strjson = strjson.replace("]","") 
#     f.write(strjson+',\n')
#     f.close()
#     return 'ok'

@app.route('/addboard',methods=['HEAD','GET','POST'])
def addboard():
    jsonData = request.get_json()
    print(jsonData["WR_ID"])
    print(jsonData["WR_BODY"])
    conn = sqlite3.connect("user_board.db", isolation_level=None)
    c = conn.cursor()
    val = [jsonData["WR_ID"],jsonData["WR_BODY"]]
    c.execute("Insert INTO user_board (name,board) VALUES(?,?)",val) 
    temp_id = jsonData["WR_ID"] 
    
    workbook = xlsxwriter.Workbook("C:/Users/real1/OneDrive/Desktop/Tempus_flask/chat/"+str(jsonData["WR_BODY"])+'.xlsx')
    worksheet = workbook.add_worksheet()
    
    worksheet.set_column('A:A', 40) # A 열의 너비를 40으로 설정
    worksheet.set_row(0,18) # A열의 높이를 18로 설정
    worksheet.set_column('B:B', 12) # B 열의 너비를 12로 설정
    worksheet.set_column('C:C', 60) # C 열의 너비를 60으로 설정

    worksheet.write(0, 0, "ID")
    worksheet.write(0, 1, '내용')
    worksheet.write(0, 2, '타입')
    worksheet.write(0, 3, 'date')
    worksheet.write(0, 4, 'WR_PNAME')
    worksheet.write(0, 5, 'WR_PRICE')
    worksheet.write(0, 6, 'WR_TAG')
    worksheet.write(0, 7, 'WR_MEMO')
    workbook.close()
    print(jsonData)

    return 'ok'

@app.route('/board',methods=['HEAD','GET','POST'])
def send_board():
    jsonData = request.get_json()
    # print(jsonData["email"])
    # print(jsonData["WR_ID"])
    # print(jsonData["WR_TYPE"])
    # print(jsonData["WR_DATE"])
    # print(jsonData["WR_BODY"])
    try:
        conn = sqlite3.connect("user_board.db", isolation_level=None)
        c = conn.cursor()
        c.execute("SELECT board FROM user_board WHERE name=:id",{"id":jsonData["email"]})
        
        fw = open("temp.txt","w")
        for row in c.fetchall():
            row_str = str(row)
            # print(row_str)
            row_str=row_str.replace("(","")
            row_str=row_str.replace(")","")
            row_str=row_str.replace(",","")
            row_str=row_str.replace("'","")
            data = { 
                'text': row_str 
                }
            fw.write(str(data)+',\n')
        fw.close()
        fw = open("temp.txt","r")
        load = fw.read()
        fw.close()
        board_list = '[' + str(load) + ']'
        # print(board_list)
        return board_list
    except:
        print("boarderrer")

@app.route('/imgupload',methods=['GET','POST'])
def imgupload():

    file = request.files["uploadedfile"]
    # print(temp_id)
    file.save("C:/Users/real1/OneDrive/Desktop/Tempus_flask/img/"+secure_filename(file.filename))

    return "업로드 성공"

@app.route('/imgdownload',methods=['GET','POST'])
def imgsend():
    fileName = request.get_json()
    try:
        if request.method == 'POST':
            fw = open("temp_img.txt","w")
            fw.write(fileName["name"])
            fw.close()
            return "ok"
        elif request.method == 'GET':
            fw = open("temp_img.txt","r")
            load = fw.read()
            fw.close()
            return send_file("C:/Users/real1/OneDrive/Desktop/Tempus_flask/img/"+load+".jpg",mimetype='image/jpg')
        # return send_file("C:/Users/real1/OneDrive/Desktop/Tempus_flask/img/"+str(fileName["name"])+".jpg",mimetype='image/jpg')
        # return send_file("C:/Users/real1/OneDrive/Desktop/Tempus_flask/img/test1234.jpg",mimetype='image/jpg')
    except:
        print("imgerror")
        return "errer"
    # return send_file("C:/Users/real1/OneDrive/Desktop/Tempus_flask/img/"+fileName["name"]+".jpg",mimetype='image/jpg')
    # return send_file("C:/Users/real1/OneDrive/Desktop/Tempus_flask/img/test1234.jpg",mimetype='image/jpg')
    
@app.route('/login',methods=['HEAD','GET','POST'])
def login():
    jsonData = request.get_json()
    auth = 'error'
    try:
        conn = sqlite3.connect("Tempus.db", isolation_level=None)
        c = conn.cursor()
        #print(val)
        param1=jsonData["email"]
        param2=jsonData["password"]
        #c.execute("SELECT * FROM user")
        c.execute("SELECT * FROM user WHERE email=:id and password=:pw",{"id":jsonData["email"],"pw":jsonData["password"]})
        #c.execute("SELECT * FROM user WHERE email=? and password=?",param1,param2)
        if c.fetchall()==[]:
            auth='error'
            print("null")
        else:
            auth='ok'
            print("ok")
        
    except:
        # c.execute("SELECT * FROM user")
        # print(c.fetchall())
        auth='error'
        print("error")
    #print(jsonData)
    # json_loginstring = jsonData['email']
    #print(jsonData["email"])
    return auth

@app.route('/signup',methods=['HEAD','GET','POST'])
def signup():
    jsonData = request.get_json()
    problem = "ok"
    #print(jsonData)
    try:
        conn = sqlite3.connect("Tempus.db", isolation_level=None)
        c = conn.cursor()
        val = [jsonData["email"],jsonData["name"],jsonData["pnum"],jsonData["address"],jsonData["password"]]
        #print(val)
        c.execute("INSERT INTO user \
            VALUES(?,?,?,?,?)", val)
        c.execute("SELECT * FROM user")
        print(c.fetchall())  
    except:
        problem = "error"
        print("error")
    # json_loginstring = jsonData['email']
    # print(jsonData["email"])
    return problem

@app.route('/addPost',methods=['HEAD','GET','POST'])
def addpost():
    jsonData = request.get_json()
    print(jsonData)
    now = datetime.now()
    curtime = str(now.month) + "월" + str(now.day) + "일 " + str(now.hour) + ":" + str(now.minute)
    wb = openpyxl.load_workbook(r"C:/Users/real1/OneDrive/Desktop/Tempus_flask/chat/"+str(jsonData["GROUP"])+'.xlsx')
    sheet = wb.active
    last_row = sheet.max_row
    if str(jsonData["WR_TYPE"]) == "1":
        sheet["A"+str(last_row+1)]=str(jsonData["WR_ID"])
        sheet["B"+str(last_row+1)]=str(jsonData["WR_BODY"])
        sheet["C"+str(last_row+1)]=str(jsonData["WR_TYPE"])
        sheet["D"+str(last_row+1)]=str(curtime)
    elif str(jsonData["WR_TYPE"]) == "2":
        sheet["A"+str(last_row+1)]=str(jsonData["WR_ID"])
        sheet["C"+str(last_row+1)]=str(jsonData["WR_TYPE"])
        sheet["D"+str(last_row+1)]=str(curtime)
        sheet["E"+str(last_row+1)]=str(jsonData["WR_PNAME"])
        sheet["F"+str(last_row+1)]=str(jsonData["WR_PRICE"])
        sheet["G"+str(last_row+1)]=str(jsonData["WR_TAG"])
        sheet["H"+str(last_row+1)]=str(jsonData["WR_MEMO"])
    wb.save("C:/Users/real1/OneDrive/Desktop/Tempus_flask/chat/"+str(jsonData["GROUP"])+'.xlsx')
    
    print(last_row)
    wb.close()

    
    # worksheet.write(excel_row, 0, jsonData["WR_ID"])
    # worksheet.write(excel_row, 1, jsonData["WR_BODY"])
    # worksheet.write(excel_row, 2, jsonData["WR_TYPE"])
    # worksheet.write(excel_row, 3, curtime)
    # workbook.close()

    return 'ok'

@app.route('/imgupload_Post',methods=['GET','POST'])
def imgupload_post():
    file = request.files["uploadedfile"]
    file.save("C:/Users/real1/OneDrive/Desktop/Tempus_flask/img_post/"+secure_filename(file.filename)+'.jpg')
    # img = Image("C:/Users/real1/OneDrive/Desktop/Tempus_flask/img_post/test.jpg")
    img = XLImage("C:/Users/real1/OneDrive/Desktop/Tempus_flask/img_post/"+secure_filename(file.filename)+'.jpg')
    wb = openpyxl.load_workbook("C:/Users/real1/OneDrive/Desktop/Tempus_flask/chat/"+secure_filename(file.filename)+".xlsx")
    sheet = wb.active
    last_row = sheet.max_row
    sheet.add_image(img,"I"+str(last_row))
    wb.save(filename="C:/Users/real1/OneDrive/Desktop/Tempus_flask/chat/"+secure_filename(file.filename)+".xlsx")
    wb.close()
    # print(temp_id)
    

    return "업로드 성공"

@app.route('/Post',methods=['HEAD','GET','POST'])
def post():
    jsonData = request.get_json()
    # print(jsonData)
    excel_path = 'C:/Users/real1/OneDrive/Desktop/Tempus_flask/chat/'+str(jsonData["WR_GROUP"])+'.xlsx'
    # print(excel_path)
    wb = openpyxl.load_workbook(excel_path)
    sh = wb.active
    last_row = sh.max_row
    fw = open('C:/Users/real1/OneDrive/Desktop/Tempus_flask/chat/'+str(jsonData["WR_GROUP"])+'.txt',"w")
    
    for rownum in range(1,last_row):
        # row_str = str(rownum)
        # print(row_str)
        data = { 
            'ID': str(sh["A"+str(rownum+1)].value),
            'BODY': str(sh["B"+str(rownum+1)].value), 
            'TYPE': str(sh["C"+str(rownum+1)].value),
            'DATE': str(sh["D"+str(rownum+1)].value),
            'WR_PNAME': str(sh["E"+str(rownum+1)].value),
            'WR_PRICE': str(sh["F"+str(rownum+1)].value),
            'WR_TAG': str(sh["G"+str(rownum+1)].value),
            'WR_MEMO':  str(sh["H"+str(rownum+1)].value)
            }
        fw.write(str(data)+',\n')
    fw.close()
    fw = open('C:/Users/real1/OneDrive/Desktop/Tempus_flask/chat/'+str(jsonData["WR_GROUP"])+'.txt',"r")
    load = fw.read()
    fw.close()
    wb.close()
    post_list = '[' + str(load) + ']'
    # data_list = []

    # for rownum in range(1, sh.nrows):
    #     data = OrderedDict()
    #     row_values = sh.row_values(rownum)
    #     data['ID'] = row_values[0]
    #     data['내용'] = row_values[1]
    #     data['타입'] = row_values[2]
    #     data['date'] = row_values[3]
    #     data_list.append(data)

    # j = json.dumps(data_list, ensure_ascii=False)

    # with open('C:/Users/real1/OneDrive/Desktop/Tempus_flask/chat/'+str(jsonData["WR_GROUP"])+'.json', 'w+') as f:
    #     f.write(j)
    return post_list
    

@app.route('/productRecommendation',methods=['HEAD','GET','POST'])
def productRecommendation():
    Data = request.get_json()
    print(Data["prdname"])
    prdctRcmnd(str(Data["prdname"]))
    excel_path = "C:/Users/real1/OneDrive/Desktop/Tempus_flask/xlsx/"+str(Data["prdname"])+"_sort.xlsx"
    # print(excel_path)
    wb = openpyxl.load_workbook(excel_path)
    sh = wb.active
    last_row = sh.max_row
    fw = open("C:/Users/real1/OneDrive/Desktop/Tempus_flask/xlsx/"+str(Data["prdname"])+'_sort.txt',"w")
    for rownum in range(1,last_row):
        # row_str = str(rownum)
        # print(row_str)

        data = { 
            'name': str(sh["A"+str(rownum+1)].value),
            'price': str(sh["B"+str(rownum+1)].value), 
            'link': str(sh["c"+str(rownum+1)].value),
            'score': str(sh["F"+str(rownum+1)].value)  
            }
        fw.write(str(data)+',\n')
    fw.close()
    fw = open('C:/Users/real1/OneDrive/Desktop/Tempus_flask/xlsx/'+str(Data["prdname"])+'_sort.txt',"r")
    load = fw.read()
    fw.close()
    wb.close()
    prd_list = '[' + str(load) + ']'
    return prd_list

    


if __name__ == '__main__':
    app.secret_key = 'super secret key'
    app.config['SESSION_TYPE'] = 'filesystem'
    app.run(host="192.168.0.3",debug=True)

# board 액티비티 - 현재는 게시판명을 제외한 나머지만 서버에 전달하지만, 게시판명도 게시판을 구분하기 위해 서버에 전달이 필요할듯
# (현재 게시판명은 board main에서 board, content 액티비티에 intent로 전달해서 사용 중)
# - 이하 변수들은 모두 board, content 액티비티에서 사용
# - WR_ID : 유저를 구분할 유저번호 또는 닉네임
# - WR_TYPE : 일반/지출목록 형식 구분
# - WR_DATE : 작성일자
# - WR_BODY : 글 내용
# - GROUP : 게시판명
# - 게시글 번호(현재는 추가 안되어있으나 추후 검색 기능을 추가했을 때 필요할 예정)


# content 액티비티에서만 사용하는 변수 - 모두 서버와 통신 필요
# - 댓글 작성자명
# - 댓글 내용
# - 댓글 작성 일자

# user_boar

# 기본키는 랜덤변수, 나머지는 email과 그룹명으로 구성된 db를 생성