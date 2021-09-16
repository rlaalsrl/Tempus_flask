from flask import Flask,request,jsonify,g,render_template
from flask import url_for, send_from_directory
import json
import os
from werkzeug.utils import secure_filename
import sqlite3
import pandas as pd
import numpy as np
from sklearn.model_selection import train_test_split
from eunjeon import Mecab
from collections import Counter
import tempfile
from tensorflow.keras.models import load_model
from tensorflow.keras.preprocessing.sequence import pad_sequences
from tensorflow.keras.preprocessing.text import Tokenizer
#from werkzeug import redirect
from flask.helpers import flash
from flask import send_file
import xlsxwriter
from datetime import datetime
from datetime import timedelta
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl import load_workbook
import openpyxl
from PIL import Image as PILImage
import xlrd
from collections import OrderedDict
import json



app = Flask(__name__)


@app.route('/')
def home():
    return "hello"
    # return render_template('C:/Users/real1/OneDrive/Desktop/Tempus_flask/upload.html')

# @app.route('/addboard',methods=['HEAD','GET','POST'])
# def addpost():
#     #print(request.is_json)
#     jsonData = request.get_json()
#     print(jsonData)
#     #print(jsonData[0]['Title'])
#     f = open("board.txt",'a')
#     strjson = str(jsonData).replace("[","")
#     strjson = strjson.replace("]","") 
#     f.write(strjson+',\n')
#     f.close()
#     return 'ok'

@app.route('/addboard',methods=['HEAD','GET','POST'])
def addboard():
    jsonData = request.get_json()
    print(jsonData["WR_ID"])
    print(jsonData["WR_BODY"])
    conn = sqlite3.connect("user_board.db", isolation_level=None)
    c = conn.cursor()
    val = [jsonData["WR_ID"],jsonData["WR_BODY"]]
    c.execute("Insert INTO user_board (name,board) VALUES(?,?)",val) 
    temp_id = jsonData["WR_ID"] 
    
    workbook = xlsxwriter.Workbook("C:/Users/real1/OneDrive/Desktop/Tempus_flask/chat/"+str(jsonData["WR_BODY"])+'.xlsx')
    worksheet = workbook.add_worksheet()
    
    worksheet.set_column('A:A', 40) # A 열의 너비를 40으로 설정
    worksheet.set_row(0,18) # A열의 높이를 18로 설정
    worksheet.set_column('B:B', 12) # B 열의 너비를 12로 설정
    worksheet.set_column('C:C', 60) # C 열의 너비를 60으로 설정

    worksheet.write(0, 0, "ID")
    worksheet.write(0, 1, '내용')
    worksheet.write(0, 2, '타입')
    worksheet.write(0, 3, 'date')
    workbook.close()
    print(jsonData)

    return 'ok'

@app.route('/board',methods=['HEAD','GET','POST'])
def send_board():
    jsonData = request.get_json()
    # print(jsonData["email"])
    # print(jsonData["WR_ID"])
    # print(jsonData["WR_TYPE"])
    # print(jsonData["WR_DATE"])
    # print(jsonData["WR_BODY"])
    try:
        conn = sqlite3.connect("user_board.db", isolation_level=None)
        c = conn.cursor()
        c.execute("SELECT board FROM user_board WHERE name=:id",{"id":jsonData["email"]})
        
        fw = open("temp.txt","w")
        for row in c.fetchall():
            row_str = str(row)
            # print(row_str)
            row_str=row_str.replace("(","")
            row_str=row_str.replace(")","")
            row_str=row_str.replace(",","")
            row_str=row_str.replace("'","")
            data = { 
                'text': row_str 
                }
            fw.write(str(data)+',\n')
        fw.close()
        fw = open("temp.txt","r")
        load = fw.read()
        fw.close()
        board_list = '[' + str(load) + ']'
        # print(board_list)
        return board_list
    except:
        print("boarderrer")

@app.route('/imgupload',methods=['GET','POST'])
def imgupload():

    file = request.files["uploadedfile"]
    # print(temp_id)
    file.save("C:/Users/real1/OneDrive/Desktop/Tempus_flask/img/"+secure_filename(file.filename))

    return "업로드 성공"

@app.route('/imgdownload',methods=['GET','POST'])
def imgsend():
    fileName = request.get_json()
    try:
        if request.method == 'POST':
            fw = open("temp_img.txt","w")
            fw.write(fileName["name"])
            fw.close()
            return "ok"
        elif request.method == 'GET':
            fw = open("temp_img.txt","r")
            load = fw.read()
            fw.close()
            return send_file("C:/Users/real1/OneDrive/Desktop/Tempus_flask/img/"+load+".jpg",mimetype='image/jpg')
        # return send_file("C:/Users/real1/OneDrive/Desktop/Tempus_flask/img/"+str(fileName["name"])+".jpg",mimetype='image/jpg')
        # return send_file("C:/Users/real1/OneDrive/Desktop/Tempus_flask/img/test1234.jpg",mimetype='image/jpg')
    except:
        print("imgerror")
        return "errer"
    # return send_file("C:/Users/real1/OneDrive/Desktop/Tempus_flask/img/"+fileName["name"]+".jpg",mimetype='image/jpg')
    # return send_file("C:/Users/real1/OneDrive/Desktop/Tempus_flask/img/test1234.jpg",mimetype='image/jpg')
    
@app.route('/login',methods=['HEAD','GET','POST'])
def login():
    jsonData = request.get_json()
    auth = 'error'
    try:
        conn = sqlite3.connect("Tempus.db", isolation_level=None)
        c = conn.cursor()
        #print(val)
        param1=jsonData["email"]
        param2=jsonData["password"]
        #c.execute("SELECT * FROM user")
        c.execute("SELECT * FROM user WHERE email=:id and password=:pw",{"id":jsonData["email"],"pw":jsonData["password"]})
        #c.execute("SELECT * FROM user WHERE email=? and password=?",param1,param2)
        if c.fetchall()==[]:
            auth='error'
            print("null")
        else:
            auth='ok'
            print("ok")
        
    except:
        # c.execute("SELECT * FROM user")
        # print(c.fetchall())
        auth='error'
        print("error")
    #print(jsonData)
    # json_loginstring = jsonData['email']
    #print(jsonData["email"])
    return auth

@app.route('/signup',methods=['HEAD','GET','POST'])
def signup():
    jsonData = request.get_json()
    problem = "ok"
    #print(jsonData)
    try:
        conn = sqlite3.connect("Tempus.db", isolation_level=None)
        c = conn.cursor()
        val = [jsonData["email"],jsonData["name"],jsonData["pnum"],jsonData["address"],jsonData["password"]]
        #print(val)
        c.execute("INSERT INTO user \
            VALUES(?,?,?,?,?)", val)
        c.execute("SELECT * FROM user")
        print(c.fetchall())  
    except:
        problem = "error"
        print("error")
    # json_loginstring = jsonData['email']
    # print(jsonData["email"])
    return problem

@app.route('/addPost',methods=['HEAD','GET','POST'])
def addpost():
    jsonData = request.get_json()
    print(jsonData)
    now = datetime.now()
    curtime = str(now.month) + "월" + str(now.day) + "일 " + str(now.hour) + ":" + str(now.minute)
    wb = openpyxl.load_workbook(r"C:/Users/real1/OneDrive/Desktop/Tempus_flask/chat/"+str(jsonData["GROUP"])+'.xlsx')
    sheet = wb.active
    last_row = sheet.max_row
    sheet["A"+str(last_row+1)]=str(jsonData["WR_ID"])
    sheet["B"+str(last_row+1)]=str(jsonData["WR_BODY"])
    sheet["C"+str(last_row+1)]=str(jsonData["WR_TYPE"])
    sheet["D"+str(last_row+1)]=str(curtime)
    wb.save("C:/Users/real1/OneDrive/Desktop/Tempus_flask/chat/"+str(jsonData["GROUP"])+'.xlsx')
    
    print(last_row)
    wb.close()

    
    # worksheet.write(excel_row, 0, jsonData["WR_ID"])
    # worksheet.write(excel_row, 1, jsonData["WR_BODY"])
    # worksheet.write(excel_row, 2, jsonData["WR_TYPE"])
    # worksheet.write(excel_row, 3, curtime)
    # workbook.close()

    return 'ok'

@app.route('/imgupload_Post',methods=['GET','POST'])
def imgupload_post():
    file = request.files["uploadedfile"]
    file.save("C:/Users/real1/OneDrive/Desktop/Tempus_flask/img_post/"+secure_filename(file.filename)+'.jpg')
    # img = Image("C:/Users/real1/OneDrive/Desktop/Tempus_flask/img_post/test.jpg")
    img = XLImage("C:/Users/real1/OneDrive/Desktop/Tempus_flask/img_post/"+secure_filename(file.filename)+'.jpg')
    wb = openpyxl.load_workbook("C:/Users/real1/OneDrive/Desktop/Tempus_flask/chat/"+secure_filename(file.filename)+".xlsx")
    sheet = wb.active
    last_row = sheet.max_row
    sheet.add_image(img,"E"+str(last_row))
    wb.save(filename="C:/Users/real1/OneDrive/Desktop/Tempus_flask/chat/"+secure_filename(file.filename)+".xlsx")
    wb.close()
    # print(temp_id)
    

    return "업로드 성공"

@app.route('/Post',methods=['HEAD','GET','POST'])
def post():
    jsonData = request.get_json()
    # print(jsonData)
    excel_path = 'C:/Users/real1/OneDrive/Desktop/Tempus_flask/chat/'+str(jsonData["WR_GROUP"])+'.xlsx'
    # print(excel_path)
    wb = openpyxl.load_workbook(excel_path)
    sh = wb.active
    last_row = sh.max_row
    fw = open('C:/Users/real1/OneDrive/Desktop/Tempus_flask/chat/'+str(jsonData["WR_GROUP"])+'.txt',"w")
    for rownum in range(1,last_row):
        # row_str = str(rownum)
        # print(row_str)

        data = { 
            'ID': str(sh["A"+str(rownum+1)].value),
            'BODY': str(sh["B"+str(rownum+1)].value), 
            'TPYE': str(sh["c"+str(rownum+1)].value),
            'DATE': str(sh["D"+str(rownum+1)].value)  
            }
        fw.write(str(data)+',\n')
    fw.close()
    fw = open('C:/Users/real1/OneDrive/Desktop/Tempus_flask/chat/'+str(jsonData["WR_GROUP"])+'.txt',"r")
    load = fw.read()
    fw.close()
    wb.close()
    post_list = '[' + str(load) + ']'
    # data_list = []

    # for rownum in range(1, sh.nrows):
    #     data = OrderedDict()
    #     row_values = sh.row_values(rownum)
    #     data['ID'] = row_values[0]
    #     data['내용'] = row_values[1]
    #     data['타입'] = row_values[2]
    #     data['date'] = row_values[3]
    #     data_list.append(data)

    # j = json.dumps(data_list, ensure_ascii=False)

    # with open('C:/Users/real1/OneDrive/Desktop/Tempus_flask/chat/'+str(jsonData["WR_GROUP"])+'.json', 'w+') as f:
    #     f.write(j)
    return post_list
    


if __name__ == '__main__':
    app.secret_key = 'super secret key'
    app.config['SESSION_TYPE'] = 'filesystem'
    app.run(host="192.168.0.3",debug=True)
